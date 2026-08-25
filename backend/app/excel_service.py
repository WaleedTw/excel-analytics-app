import math
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from app.config import MAX_COLUMNS, MAX_ROWS, MAX_SHEETS, SAMPLE_DIR, UPLOAD_DIR
from app.data_cleaning import coerce_numeric_series, is_numeric_like
from app.schemas import ColumnProfile, QualityReport, SheetInfo, WorkbookInfo

ALLOWED_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
    "application/zip",
}
SAFE_NAME = re.compile(r"^[\w\-. ]+$", re.UNICODE)
FORMULA_PREFIXES = ("=", "+", "-", "@")

DATE_TERMS = {"التاريخ", "تاريخ", "date", "day", "اليوم"}
TIME_DIMENSION_TERMS = {
    "year", "quarter", "month", "week", "السنة", "السنه", "العام",
    "الربع", "الشهر", "الأسبوع", "الاسبوع",
}
MEASURE_TERMS = {
    "الكمية", "الإيرادات", "الايرادات", "التكلفة", "الربح", "المبيعات",
    "السعر", "سعر", "المبلغ", "إجمالي", "اجمالي", "الإجمالي", "الاجمالي",
    "النسبة", "نسبة", "quantity", "revenue", "cost",
    "profit", "sales", "price", "amount", "value", "percentage", "percent",
    "rate", "margin",
}
DIMENSION_TERMS = {
    "المدينة", "المنطقة", "المنتج", "الفئة", "مندوب المبيعات", "المندوب",
    "الشركة", "الفرع", "city", "region", "product", "category", "salesperson",
    "العميل", "اسم العميل", "حالة الطلب", "الحالة", "segment", "company",
    "customer", "status", "branch", "brand", "channel", "country",
}
IDENTIFIER_TERMS = {"id", "المعرف", "معرف", "رقم", "code", "رمز"}


class FileValidationError(ValueError):
    pass


def _path_is_inside(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def _normalize_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, np.generic):
        return value.item()
    return value


def inspect_xlsx(path: Path, original_name: str, mime_type: str, file_id: str) -> WorkbookInfo:
    if not _path_is_inside(path, UPLOAD_DIR) and not _path_is_inside(path, SAMPLE_DIR):
        raise FileValidationError("مسار الملف خارج نطاق المشروع الآمن.")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise FileValidationError("ملف Excel تالف أو غير قابل للقراءة.") from exc
    if len(workbook.sheetnames) > MAX_SHEETS:
        workbook.close()
        raise FileValidationError(f"عدد أوراق العمل يتجاوز الحد المسموح ({MAX_SHEETS}).")
    sheets: list[SheetInfo] = []
    for worksheet in workbook.worksheets:
        dimension = worksheet.calculate_dimension(force=True)
        min_col, min_row, max_col, max_row = range_boundaries(dimension)
        rows = max_row if dimension != "A1:A1" or worksheet["A1"].value is not None else 0
        columns = max_col if rows else 0
        if rows > MAX_ROWS or columns > MAX_COLUMNS:
            workbook.close()
            raise FileValidationError("حجم ورقة العمل يتجاوز حدود الصفوف أو العواميد الآمنة.")
        sheets.append(SheetInfo(name=worksheet.title, rows=rows, columns=columns, has_data=rows > 1 and columns > 0))
    workbook.close()
    if not any(sheet.has_data for sheet in sheets):
        raise FileValidationError("لا يحتوي الملف على بيانات قابلة للتحليل.")
    return WorkbookInfo(
        file_id=file_id,
        original_name=original_name,
        safe_name=path.name,
        size_bytes=path.stat().st_size,
        mime_type=mime_type,
        sheets=sheets,
        created_at=datetime.now(timezone.utc),
    )


def read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    try:
        frame = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError as exc:
        raise FileValidationError("ورقة العمل المطلوبة غير موجودة.") from exc
    # pandas/openpyxl reads cached formula results. Some valid workbooks (for
    # example files generated without Excel recalculation) contain formulas but
    # no cached values, which otherwise makes a populated column look empty.
    # Preserve only those formula strings for the conservative cleaning stage.
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        worksheet = workbook[sheet_name]
        formula_cells: list[tuple[int, int, str]] = []
        for row_index, values in enumerate(
            worksheet.iter_rows(min_row=2, max_row=len(frame) + 1, max_col=len(frame.columns), values_only=True)
        ):
            for column_index, value in enumerate(values):
                if isinstance(value, str) and value.startswith("=") and pd.isna(frame.iat[row_index, column_index]):
                    formula_cells.append((row_index, column_index, value))
        for column_index in {item[1] for item in formula_cells}:
            frame[frame.columns[column_index]] = frame.iloc[:, column_index].astype("object")
        for row_index, column_index, value in formula_cells:
            frame.iat[row_index, column_index] = value
    finally:
        workbook.close()
    frame.columns = [str(column).strip() or f"عامود_{index + 1}" for index, column in enumerate(frame.columns)]
    # Excel cells containing only spaces look filled to pandas even though they are
    # empty to the user. Normalize them once so preview and quality counts agree.
    object_columns = frame.select_dtypes(include=["object", "string"]).columns
    if len(object_columns):
        frame[object_columns] = frame[object_columns].replace(r"^\s*$", pd.NA, regex=True)
    return frame


def infer_columns(frame: pd.DataFrame, mapping: dict[str, str] | None = None) -> list[ColumnProfile]:
    mapping = mapping or {}
    profiles: list[ColumnProfile] = []
    for column in frame.columns:
        series = frame[column]
        lower = column.strip().lower()
        normalized = re.sub(r"[_\-]+", " ", lower)
        words = set(normalized.split())
        mapped = mapping.get(column, "").lower()
        is_date_name = lower in DATE_TERMS or any(term in words for term in DATE_TERMS) or "تاريخ" in lower
        is_time_dimension = lower in TIME_DIMENSION_TERMS or any(term in words for term in TIME_DIMENSION_TERMS)
        is_dimension_name = lower in DIMENSION_TERMS or any(term in words for term in DIMENSION_TERMS)
        is_measure_name = lower in MEASURE_TERMS or any(term in words for term in MEASURE_TERMS)
        is_identifier_name = lower in IDENTIFIER_TERMS or any(term in words for term in IDENTIFIER_TERMS) or lower.endswith("_id") or lower.endswith(" id")
        numeric_like = is_numeric_like(series) if not pd.api.types.is_numeric_dtype(series) else True
        empty_column = series.dropna().empty
        if empty_column:
            inferred_type, role = "unknown", "unknown"
        elif mapped in {"date", "تاريخ"}:
            inferred_type, role = "date", "date"
        elif mapped in {"dimension", "بُعد", "بعد"}:
            inferred_type, role = "category", "dimension"
        elif mapped in {"identifier", "معرف"}:
            inferred_type, role = "text", "identifier"
        elif mapped in {"measure", "مقياس"}:
            inferred_type, role = "number", "measure"
        elif pd.api.types.is_datetime64_any_dtype(series) or is_date_name:
            inferred_type, role = "date", "date"
        elif is_identifier_name:
            inferred_type, role = "text", "identifier"
        elif is_time_dimension or is_dimension_name:
            inferred_type, role = "category", "dimension"
        elif is_measure_name or numeric_like:
            inferred_type, role = "number", "measure"
        elif series.nunique(dropna=True) <= max(20, len(series) * 0.2):
            inferred_type, role = "category", "unknown"
        else:
            inferred_type, role = "text", "unknown"
        ambiguous = role == "unknown" and column not in mapping and not empty_column
        profiles.append(
            ColumnProfile(
                name=column,
                inferred_type=inferred_type,
                semantic_role=role,
                null_count=int(series.isna().sum()),
                unique_count=int(series.nunique(dropna=True)),
                sample_values=[_normalize_value(v) for v in series.dropna().head(3).tolist()],
                ambiguous=ambiguous,
                reason=(
                    "العامود فارغ بالكامل وسيُستبعد من نسخة التحليل."
                    if empty_column else
                    "لم يتطابق الاسم مع قاموس الدلالات المحلي."
                    if ambiguous else
                    "تم الاستدلال من الاسم ونوع القيم."
                ),
            )
        )
    return profiles


def profile_quality(frame: pd.DataFrame, profiles: list[ColumnProfile]) -> QualityReport:
    total_cells = max(frame.shape[0] * frame.shape[1], 1)
    missing = int(frame.isna().sum().sum())
    duplicates = int(frame.duplicated().sum())
    text_frame = frame.select_dtypes(include=["object", "string"])
    formula_like = int(sum(series.astype(str).str.startswith(FORMULA_PREFIXES).sum() for _, series in text_frame.items()))
    invalid = 0
    outliers = 0
    for profile in profiles:
        series = frame[profile.name]
        if profile.semantic_role == "measure":
            numeric, _, invalid_count = coerce_numeric_series(series)
            invalid += invalid_count
            clean = numeric.dropna()
            if len(clean) >= 4:
                q1, q3 = clean.quantile([0.25, 0.75])
                spread = q3 - q1
                if spread > 0:
                    outliers += int(((clean < q1 - 1.5 * spread) | (clean > q3 + 1.5 * spread)).sum())
        if profile.semantic_role == "date":
            parsed = pd.to_datetime(series, errors="coerce")
            invalid += int((series.notna() & parsed.isna()).sum())
    penalty = min(100, round((missing / total_cells) * 55 + (duplicates / max(len(frame), 1)) * 20 + invalid * 2 + outliers))
    notes = []
    if missing: notes.append(f"تم رصد {missing} خلية ناقصة.")
    if duplicates: notes.append(f"تم رصد {duplicates} صفوف مكررة.")
    if invalid: notes.append(f"تم رصد {invalid} قيمة لا تطابق النوع المتوقع.")
    if outliers: notes.append(f"تم رصد {outliers} قيمة شاذة إحصائيًا.")
    if formula_like: notes.append(f"تم تحييد {formula_like} قيمة نصية تشبه الصيغ لحماية العرض.")
    if not notes: notes.append("لم تُرصد مشكلات جوهرية في جودة البيانات.")
    return QualityReport(
        row_count=len(frame), column_count=len(frame.columns), missing_cells=missing,
        missing_rate=round(missing / total_cells, 4), duplicate_rows=duplicates,
        invalid_values=invalid, outlier_count=outliers, formula_like_cells=formula_like,
        score=max(0, 100 - penalty), notes=notes,
    )